# backend/app/utils/excel.py
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict
from io import BytesIO


def generate_import_template() -> BytesIO:
    """生成导入模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "客户导入模板"

    # 设置表头
    headers = [
        "客户名称*",
        "客户编码",
        "行业",
        "负责销售 ID*",
        "价值等级",
        "年消费金额",
        "联系人",
        "联系电话",
        "联系邮箱",
        "地址",
        "备注",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center")

    # 设置列宽
    column_widths = [20, 15, 15, 15, 10, 15, 15, 15, 25, 40, 40]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width

    # 添加示例数据
    ws.append(
        [
            "XX 科技公司",
            "CUST001",
            "科技",
            "1",
            "A",
            "500000.00",
            "张三",
            "13800138000",
            "zhangsan@example.com",
            "北京市朝阳区",
            "示例备注",
        ]
    )

    # 保存到 BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def parse_import_excel(file_content: bytes) -> List[Dict]:
    """解析导入的 Excel 文件"""
    wb = load_workbook(BytesIO(file_content))
    ws = wb.active

    customers = []
    errors = []

    # 跳过表头,从第 2 行开始
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not any(row):  # 空行跳过
            continue

        # 解析数据
        customer_data = {
            "name": row[0] if len(row) > 0 else None,
            "code": row[1] if len(row) > 1 else None,
            "industry": row[2] if len(row) > 2 else None,
            "sales_rep_id": row[3] if len(row) > 3 else None,
            "tier_level": row[4] if len(row) > 4 else None,
            "annual_consumption": row[5] if len(row) > 5 else None,
            "contact_person": row[6] if len(row) > 6 else None,
            "contact_phone": row[7] if len(row) > 7 else None,
            "contact_email": row[8] if len(row) > 8 else None,
            "address": row[9] if len(row) > 9 else None,
            "remark": row[10] if len(row) > 10 else None,
        }

        # 验证必填字段
        if not customer_data["name"]:
            errors.append(
                {
                    "row": row_idx,
                    "field": "客户名称",
                    "message": "客户名称不能为空",
                }
            )
            continue

        if not customer_data["sales_rep_id"]:
            errors.append(
                {
                    "row": row_idx,
                    "field": "负责销售 ID",
                    "message": "负责销售 ID 不能为空",
                }
            )
            continue

        customers.append({"data": customer_data, "row": row_idx})

    return {"customers": customers, "errors": errors}


def generate_export_excel(customers: List[Dict]) -> BytesIO:
    """生成导出的 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "客户数据"

    # 设置表头
    headers = [
        "客户 ID",
        "客户名称",
        "客户编码",
        "行业",
        "负责销售 ID",
        "价值等级",
        "年消费金额",
        "状态",
        "联系人",
        "联系电话",
        "联系邮箱",
        "地址",
        "备注",
        "创建时间",
        "更新时间",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center")

    # 添加客户数据
    for customer in customers:
        ws.append(
            [
                customer.get("id"),
                customer.get("name"),
                customer.get("code"),
                customer.get("industry"),
                customer.get("sales_rep_id"),
                customer.get("tier_level"),
                customer.get("annual_consumption"),
                customer.get("status"),
                customer.get("contact_person"),
                customer.get("contact_phone"),
                customer.get("contact_email"),
                customer.get("address"),
                customer.get("remark"),
                customer.get("created_at"),
                customer.get("updated_at"),
            ]
        )

    # 保存到 BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
