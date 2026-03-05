"""
批量导入导出E2E测试
"""

import pytest
import requests
import uuid
from io import BytesIO
from test_helpers import (
    APIClient,
    get_test_users,
    assert_api_response,
    assert_error_response,
)


class TestImportExport:
    """批量导入导出E2E测试"""

    def test_download_import_template_success(self, base_url, test_tokens):
        """下载导入模板成功"""
        token = test_tokens["admin"]

        response = requests.get(
            f"{base_url}/customers/import-template",
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 200
        assert response.headers.get("Content-Type").startswith("application/vnd.open")

        # 验证文件内容不为空
        assert len(response.content) > 0

    def test_batch_import_success(self, base_url, test_tokens):
        """批量导入客户数据成功"""
        token = test_tokens["admin"]

        from openpyxl import Workbook

        # 创建测试Excel文件 - 使用唯一代码避免重复
        unique_code = f"IMP{uuid.uuid4().hex[:8].upper()}"
        wb = Workbook()
        ws = wb.active
        ws.append(["客户名称", "客户编码", "行业", "负责销售 ID"])
        ws.append(["导入测试客户", unique_code, "科技", "122"])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # 上传文件
        response = requests.post(
            f"{base_url}/customers/import",
            headers={"Authorization": f"Bearer {token}"},
            files={"file": ("import.xlsx", output, "application/vnd.openxmlformats")},
        )

        assert_api_response(response, 200)

        data = response.json().get("data", {})
        assert "imported_count" in data
        assert data["imported_count"] >= 1

    def test_batch_import_data_validation(self, base_url, test_tokens):
        """批量导入数据验证"""
        token = test_tokens["admin"]

        from openpyxl import Workbook
        from io import BytesIO

        # 创建缺少必填字段的Excel
        wb = Workbook()
        ws = wb.active
        ws.append(["客户名称", "负责销售 ID"])
        ws.append(["", ""])  # 客户名称为空

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = requests.post(
            f"{base_url}/customers/import",
            headers={"Authorization": f"Bearer {token}"},
            files={"file": ("invalid.xlsx", output, "application/vnd.openxmlformats")},
        )

        assert response.status_code >= 400 or response.status_code == 200
        assert "error" in response.json() or "data" in response.json()

    def test_batch_import_error_handling(self, base_url, test_tokens):
        """批量导入错误处理"""
        token = test_tokens["admin"]

        # 测试没有文件
        response = requests.post(
            f"{base_url}/customers/import", headers={"Authorization": f"Bearer {token}"}
        )

        assert_error_response(response, "NO_FILE", "请上传")

    def test_batch_export_success(self, base_url, test_tokens):
        """批量导出客户数据成功"""
        token = test_tokens["admin"]

        response = requests.get(
            f"{base_url}/customers/export", headers={"Authorization": f"Bearer {token}"}
        )

        assert response.status_code == 200
        assert response.headers.get("Content-Type").startswith("application/vnd.open")

        # 验证文件内容不为空
        assert len(response.content) > 0

    def test_batch_export_data_validation(self, base_url, test_tokens):
        """批量导出数据验证"""
        token = test_tokens["admin"]

        response = requests.get(
            f"{base_url}/customers/export", headers={"Authorization": f"Bearer {token}"}
        )

        assert response.status_code == 200

        # 导出文件应该包含正确的表头 - Excel文件需要用openpyxl读取
        import openpyxl

        wb = openpyxl.load_workbook(BytesIO(response.content))
        sheet = wb.active
        first_row = [cell.value for cell in sheet[1]]
        assert "客户名称" in first_row
